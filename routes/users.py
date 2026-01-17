
from fastapi import APIRouter, HTTPException, status, Depends, Request, Body, WebSocket, WebSocketDisconnect
from db.dbconn import users_collections, groups, tasks, completedtasks, fs, comments, chat_read_state, telegram_users
from db.hash import Hash
from typing import Dict
from jose import jwt
from fastapi.encoders import jsonable_encoder
import os
from shemas.users import UserLogin, UserRegister, DeleteUserRequest, GroupCreateRequest, DeleteGroupRequest, UserEdit, GroupEdit, Task, TaskTime,TaskTimeCancel, TaskEdit, GroupCreateRequest2, GroupCreateRequest3, TaskRequest, QuestionTaskRequest, ChatReadRequest
from middelware.auth import auth_middleware_status_return, verify_admin_token, auth_middleware_phone_return
from bson import ObjectId
from io import BytesIO
from datetime import datetime
from pymongo.errors import PyMongoError
from datetime import datetime, timedelta
import calendar
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from jwt.exceptions import ExpiredSignatureError, InvalidTokenError
from urllib.parse import unquote
from fastapi import Form, File, UploadFile, Request, Depends
from typing import List
import json
import gridfs
from telegramfiles.startpage import bot

user_app = APIRouter()  

@user_app.post("/login")
async def login_user(user: UserLogin):
    """
    Authenticates a user and returns a JWT token upon successful login.
    
    Args:
        user (UserLogin): User login credentials including phone and password
        
    Returns:
        dict: JWT token if authentication is successful
        
    Raises:
        HTTPException: 400 if user not found or credentials are invalid
        
    Example Request:
        POST /login
        {
            "phone": "+380123456789",
            "password": "securepassword123"
        }
        
    Example Response:
        {
            "token": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9..."
        }
    """
    try:
        found_user = await users_collections.find_one({"phone": user.phone})
        if not found_user:
            return HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User not found"
            )
        if not Hash.verify(user.password, found_user["password"]):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid credentials"
            )
        token = jwt.encode(
            {
                'sub': found_user["phone"],
                'status': found_user['status'],
                'exp': datetime.utcnow() + timedelta(hours=24)  
            },
            os.getenv("SecretJwt"),
            algorithm='HS256'
        )
        return {"token": token}
    except Exception as e:
        print(e)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Internal server error"
        )

@user_app.get("/get_status/{token}")
async def login_user(token:str):
    """
    Decodes a JWT token and returns the user's status.
    
    Args:
        token (str): JWT token to decode
        
    Returns:
        str: User status extracted from the token
        
    Raises:
        HTTPException: If token is invalid
        
    Example Request:
        GET /get_status/eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9...
        
    Example Response:
        "admin"
    """
    try:
        payload = jwt.decode(token, os.getenv("SecretJwt"), algorithms=["HS256"])
        user_status = payload.get("status")
        if not user_status:
            raise HTTPException(status_code=400, detail="User status not found in token")
        return str(user_status)
    except ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    except Exception as e:
        raise HTTPException(status_code=500, detail="Internal server error during token decoding")

@user_app.post("/register", dependencies=[Depends(verify_admin_token)])
async def register_user(request: Request, user: UserRegister):
    """
    Registers a new user in the system (admin-only).
    
    Args:
        user (UserRegister): User registration details including name, phone, password, status
        
    Returns:
        dict: Registration status
        
    Raises:
        HTTPException: 400 if user already exists or database error occurs
        
    Example Request:
        POST /register
        {
            "name": "John Doe",
            "phone": "+380987654321",
            "password": "newpassword123",
            "status": "user",
            "telegramName": "johndoe"
        }
        
    Example Response:
        {"status": "Ok"}
    """
    existing_user = await users_collections.find_one({"phone": user.phone})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="User already exists"
        )

    hashed_password = Hash.bcrypt(user.password)
    user.password = hashed_password

    try:
        await users_collections.insert_one(user.dict())
    except PyMongoError as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Database error occurred while registering the user"
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Unexpected error occurred"
        )

    return {"status": "User registered successfully"}

@user_app.get("/task/{task_id}")
async def get_task(task_id: str):
    """
    Retrieves a specific task by its ID.
    
    Args:
        task_id (str): MongoDB ObjectId of the task
        
    Returns:
        dict: Task details
        
    Raises:
        HTTPException: 400 for invalid ID format, 404 if task not found
        
    Example Request:
        GET /task/507f1f77bcf86cd799439011
        
    Example Response:
        {
            "_id": "507f1f77bcf86cd799439011",
            "title": "Complete project",
            "description": "Finish the API documentation",
            ...
        }
    """
    try:
        obj_id = ObjectId(task_id)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Невірний формат ID"
        )

    try:
        task = await tasks.find_one({"_id": obj_id})
    except PyMongoError as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Помилка при зверненні до бази даних"
        )

    if not task:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Завдання не знайдено"
        )

    task["_id"] = str(task["_id"])
    return task
    
@user_app.post("/tasks_by_group")
async def get_tasks_by_group(start_date: str, end_date: str, phone=Depends(auth_middleware_phone_return)):
    """
    Retrieves completed tasks for the authenticated user grouped by task group within a date range.
    
    Args:
        start_date (str): Start date in YYYY-MM-DD format
        end_date (str): End date in YYYY-MM-DD format
        phone (str): Authenticated user's phone (from dependency)
        
    Returns:
        dict: Tasks grouped by their group with total active minutes
        
    Raises:
        HTTPException: 400 for invalid date format
        
    Example Request:
        POST /tasks_by_group
        start_date=2023-01-01&end_date=2023-01-31
        
    Example Response:
        {
            "Group1": [
                450,  // total active minutes
                {task1...},
                {task2...}
            ],
            "Group2": [...]
        }
    """
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Неправильний формат дати. Використовуйте YYYY-MM-DD")

    query = {
        "phone": phone,
        "finish_time": {"$exists": True}
    }

    try:
        tasks_cursor = completedtasks.find(query)
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    grouped_tasks = {}

    async for task in tasks_cursor:
        try:
            finish_time = datetime.strptime(task["finish_time"], "%d.%m.%Y, %H:%M:%S")
        except (ValueError, KeyError):
            continue  

        if start_dt <= finish_time <= end_dt:
            group = task.get("group", "Без групи")
            task["_id"] = str(task["_id"])
            task["__parsed_finish_time"] = finish_time
            grouped_tasks.setdefault(group, []).append(task)

    result = {}
    for group, tasks in grouped_tasks.items():
        sorted_tasks = sorted(tasks, key=lambda x: x["__parsed_finish_time"], reverse=True)

        for t in sorted_tasks:
            del t["__parsed_finish_time"]  
            total_active_minutes = sum(t.get("active_minutes", 0) or 0 for t in sorted_tasks)
            
        result[group] = [total_active_minutes] + sorted_tasks

    return result

@user_app.post("/tasks_by_group2")
async def get_tasks_by_group(request: TaskRequest, phone=Depends(auth_middleware_phone_return)):
    """
    Retrieves completed tasks for a specific group within a date range (manager-only).
    
    Args:
        request (TaskRequest): Contains start_date, end_date, and group name
        phone (str): Authenticated manager's phone (from dependency)
        
    Returns:
        list: Tasks with total active minutes as first element
        
    Raises:
        HTTPException: 403 if not group manager, 404 if group not found, 400 for invalid dates
        
    Example Request:
        POST /tasks_by_group2
        {
            "start_date": "2023-01-01",
            "end_date": "2023-01-31",
            "group": "Development"
        }
        
    Example Response:
        [
            450,  // total active minutes
            {task1 data...},
            {task2 data...}
        ]
    """
    try:
        phone_group = await groups.find_one({"group_name": request.group}, {"manager_phone": 1})
        if phone_group["manager_phone"] != phone:
            raise HTTPException(status_code=403, detail="У вас немає прав для виконання цієї задачі")
    except Exception:
        raise HTTPException(status_code=404, detail="Група не знайдена")
    try:
        start_dt = datetime.strptime(request.start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(request.end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Неправильний формат дати")

    query = {
        "group": request.group,
        "finish_time": {"$exists": True}
    }

    tasks_cursor = completedtasks.find(query)

    result = []
    total_active_minutes = 0  

    async for task in tasks_cursor:
        try:
            finish_time = datetime.strptime(task["finish_time"], "%d.%m.%Y, %H:%M:%S")
        except Exception:
            continue

        if start_dt <= finish_time <= end_dt:
            task["_id"] = str(task["_id"])
            active_minutes = task.get("active_minutes")
            if isinstance(active_minutes, (int, float)):
                total_active_minutes += active_minutes

            result.append(task)
    result2 = [total_active_minutes] + result
    return result2

@user_app.get("/get_users", dependencies=[Depends(verify_admin_token)])
async def get_users(request: Request):
    """
    Retrieves list of all non-admin users (admin-only).
    
    Returns:
        list: All users excluding admins
        
    Example Request:
        GET /get_users
        
    Example Response:
        [
            {
                "_id": "507f1f77bcf86cd799439011",
                "name": "John Doe",
                "phone": "+380123456789",
                "status": "manager"
            },
            ...
        ]
    """
    try:
        cursor = users_collections.find({"status": {"$ne": "admin"}})
        users_list = []

        async for user in cursor:
            user["_id"] = str(user["_id"])  
            users_list.append(user)

        return users_list

    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка доступу до бази даних")

@user_app.post("/delete_user", dependencies=[Depends(verify_admin_token)])
async def delete_user(request: Request, user: DeleteUserRequest):
    """
    Deletes a user and all associated data (admin-only).
    
    Args:
        user (DeleteUserRequest): Contains user ID and phone
        
    Returns:
        dict: Success message
        
    Raises:
        HTTPException: 400 for invalid ID, 404 if user not found
        
    Example Request:
        POST /delete_user
        {
            "id": "507f1f77bcf86cd799439011",
            "phone": "+380123456789"
        }
        
    Example Response:
        {"message": "User successfully deleted"}
    """
    try:
        if not ObjectId.is_valid(user.id):
            raise HTTPException(status_code=400, detail="Невірний формат ID")

        result = await users_collections.delete_one({"_id": ObjectId(user.id)})

        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Користувача не знайдено")

        tasks_for_delete = []
        async for group in groups.find({"manager_phone": user.phone}, {'_id': 0, 'group_name': 1}):
            tasks_for_delete.append(group['group_name'])

        await groups.delete_many({"manager_phone": user.phone})

        await groups.update_many(
            {"user_phones": {"$in": [user.phone]}},
            {"$pull": {"user_phones": user.phone}}
        )

        if tasks_for_delete:
            await tasks.delete_many({'group': {"$in": tasks_for_delete}})
        return {"message": "Користувача успішно видалено"}

    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.post("/delete_group", dependencies=[Depends(verify_admin_token)])
async def delete_group(request: Request, group: DeleteGroupRequest):
    """
    Deletes a group and all associated tasks (admin-only).
    
    Args:
        group (DeleteGroupRequest): Contains group name to delete
        
    Returns:
        dict: Success message
        
    Raises:
        HTTPException: 404 if group not found
        
    Example Request:
        POST /delete_group
        {
            "group_name": "Development"
        }
        
    Example Response:
        {"message": "Group successfully deleted"}
    """
    try:
        result = await groups.delete_one({"group_name": group.group_name})

        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Group not found")

        await tasks.delete_many({'group': group.group_name})
        
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")
    
    return {"message": "Group successfully deleted"}

@user_app.get("/get_users_add", dependencies=[Depends(verify_admin_token)])
async def get_users_add(request: Request):
    """
    Retrieves users with 'add' status (admin-only).
    
    Returns:
        list: Users filtered by add status (excluding admin/receive)
        
    Example Request:
        GET /get_users_add
        
    Example Response:
        [
            {
                "name": "Alice Smith",
                "phone": "+380987654321"
            },
            ...
        ]
    """
    try:
        cursor = users_collections.find(
            {"status": {"$nin": ["admin", "receive"]}},
            {"name": 1, "phone": 1, "_id": 0}
        )
        return [user async for user in cursor]
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.get("/get_users_receive", dependencies=[Depends(verify_admin_token)])
async def get_users_receive(request: Request):
    """
    Retrieves users with 'receive' status (admin-only).
    
    Returns:
        list: Users filtered by receive status (excluding admin/add)
        
    Example Request:
        GET /get_users_receive
        
    Example Response:
        [
            {
                "name": "Bob Johnson",
                "phone": "+380111223344"
            },
            ...
        ]
    """
    try:
        cursor = users_collections.find(
            {"status": {"$nin": ["admin", "add"]}},
            {"name": 1, "phone": 1, "_id": 0}
        )
        return [user async for user in cursor]
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.post("/create_group/", dependencies=[Depends(verify_admin_token)])
async def create_group(group: GroupCreateRequest):
    """
    Creates a new group (admin-only).
    
    Args:
        group (GroupCreateRequest): Group creation data
        
    Returns:
        dict: Success message
        
    Raises:
        HTTPException: 400 if group already exists
        
    Example Request:
        POST /create_group
        {
            "group_name": "Development",
            "manager_phone": "+380123456789",
            "user_phones": ["+380987654321", "+380111223344"]
        }
        
    Example Response:
        {"message": "Group successfully created"}
    """
    try:
        existing_group = await groups.find_one({"group_name": group.group_name})
        if existing_group:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The group with this name already exists")

        group_data = {
            "group_name": group.group_name,
            "manager_phone": group.manager_phone,
            "user_phones": group.user_phones,
            "active": 1
        }
        await groups.insert_one(group_data)
        return {"message": "Group successfully created"}
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.post("/get_users_info_group")
async def get_users_info_group(
    group: GroupCreateRequest2,
    phone: str = Depends(auth_middleware_phone_return)
):
    """
    Retrieves member information for a specific group (manager-only).
    
    Args:
        group (GroupCreateRequest2): Contains group name
        phone (str): Authenticated user's phone (must be group manager)
        
    Returns:
        dict: Group members' information
        
    Raises:
        HTTPException: 403 if not manager, 404 if group not found
        
    Example Request:
        POST /get_users_info_group
        {
            "group_name": "Development"
        }
        
    Example Response:
        {
            "users": [
                {
                    "name": "Alice Smith",
                    "phone": "+380987654321",
                    "telegramName": "alice_smith"
                },
                ...
            ]
        }
    """
    try:
        group_data = await groups.find_one({"group_name": group.group_name})
        if not group_data:
            raise HTTPException(status_code=404, detail="Group not found")
        if group_data["manager_phone"] != phone:
            raise HTTPException(status_code=403, detail="You do not have permission to access this group")
        user_phones = group_data.get("user_phones", [])
        cursor = users_collections.find({"phone": {"$in": user_phones}}, {"password": 0, "_id": 0, "status": 0})
        users = await cursor.to_list(length=None)  
        return {"users": users}
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.post("/get_users_info_group2")
async def get_users_info_group2(
    groups_req: GroupCreateRequest3,
    phone: str = Depends(auth_middleware_phone_return)
):
    """
    Retrieves member information for multiple groups (manager-only).
    
    Args:
        groups_req (GroupCreateRequest3): Contains list of group names
        phone (str): Authenticated user's phone (must be manager of all groups)
        
    Returns:
        dict: Members per group
        
    Raises:
        HTTPException: 403 if not manager of any group
        
    Example Request:
        POST /get_users_info_group2
        {
            "groups_names": ["Development", "Design"]
        }
        
    Example Response:
        {
            "Development": [
                {user1 data...},
                {user2 data...}
            ],
            "Design": [...]
        }
    """
    try:
        result = {}

        for group_name in groups_req.groups_names:
            group_data = await groups.find_one({"group_name": group_name})
            
            if not group_data:
                continue 
            
            if group_data["manager_phone"] != phone:
                continue  

            user_phones = group_data.get("user_phones", [])
            
            cursor = users_collections.find(
                {"phone": {"$in": user_phones}},
                {"password": 0, "_id": 0, "status": 0}
            )
            users = await cursor.to_list(length=None)
            
            result[group_name] = users
        
        return result
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.get("/download_file/{file_id}")
async def download_file(file_id: str):
    """
    Downloads a file by its ID.
    
    Args:
        file_id (str): MongoDB GridFS file ID
        
    Returns:
        StreamingResponse: File download stream
        
    Raises:
        HTTPException: 400 for invalid ID, 404 if file not found
        
    Example Request:
        GET /download_file/507f1f77bcf86cd799439011
        
    Response:
        Binary file stream with Content-Disposition header
    """
    try:
        file_object_id = ObjectId(file_id)  
    except Exception as e:
        raise HTTPException(status_code=400, detail="Invalid file ID format")
    try:
        file_stream = await fs.open_download_stream(file_object_id)
    except gridfs.errors.NoFile:
        raise HTTPException(status_code=404, detail="File not found")
    return StreamingResponse(file_stream, media_type="image/jpeg", headers={"Content-Disposition": f"attachment;"})

@user_app.get("/get_groups/", dependencies=[Depends(verify_admin_token)])
async def get_groups(request: Request):
    """
    Retrieves list of all groups (admin-only).
    
    Returns:
        list: All groups with basic info
        
    Example Request:
        GET /get_groups/
        
    Example Response:
        [
            {
                "group_name": "Development",
                "manager_phone": "+380123456789",
                "user_phones": ["+380987654321"],
                "active": 1
            },
            ...
        ]
    """
    try:
        cursor = groups.find({}, {
            "group_name": 1,    
            "manager_phone": 1,  
            "user_phones": 1,
            "active": 1,
            "_id": 0     
        })
        return [group async for group in cursor] 
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.post("/edit_user/", dependencies=[Depends(verify_admin_token)])
async def edit_user(request: Request, user: UserEdit):
    """
    Updates user information (admin-only).
    
    Args:
        user (UserEdit): User data to update
        
    Returns:
        dict: Success message
        
    Raises:
        HTTPException: 400 if no valid fields, 404 if user not found
        
    Example Request:
        POST /edit_user/
        {
            "id": "507f1f77bcf86cd799439011",
            "name": "New Name",
            "status": "manager",
            "password": "newpassword123",
            "telegramName": "new_telegram"
        }
        
    Example Response:
        {"message": "User updated successfully"}
    """
    try:
        user_id = ObjectId(user.id)
        update_data = {}

        if user.name:
            update_data["name"] = user.name
        if user.password:
            update_data["password"] = Hash.bcrypt(user.password)
        if user.status is not None: 
            update_data["status"] = user.status
        if user.telegramName: 
            update_data["telegramName"] = user.telegramName

        if not update_data:
            raise HTTPException(status_code=400, detail="No valid fields to update")

        result = await users_collections.update_one(
            {"_id": user_id},
            {"$set": update_data}
        )

        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="User not found")

        return {"message": "User updated successfully"}
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.post("/edit_group/", dependencies=[Depends(verify_admin_token)])
async def edit_group(request: Request, user: GroupEdit):
    """
    Updates group information (admin-only).
    
    Args:
        user (GroupEdit): Group data to update
        
    Returns:
        dict: Success message
        
    Raises:
        HTTPException: 400 if no valid fields, 404 if group not found
        
    Example Request:
        POST /edit_group/
        {
            "group_name": "Development",
            "manager_phone": "+380987654321",
            "user_phones": ["+380111223344", "+380555667788"],
            "active": 1
        }
        
    Example Response:
        {"message": "Group updated successfully"}
    """
    try:
        update_data = {}

        if user.manager_phone:
            update_data["manager_phone"] = user.manager_phone
        if user.user_phones:
            update_data["user_phones"] = user.user_phones
        update_data["active"] = user.active
    
        if not update_data:
            raise HTTPException(status_code=400, detail="No valid fields to update")

        result = await groups.update_one(
            {"group_name": user.group_name},
            {"$set": update_data}
        )

        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Group not found")

        return {"message": "Group updated successfully"}
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")
  
@user_app.get("/get_my_groups_analytic")
async def login_user(request: Request, start_date: str, end_date: str, phone = Depends(auth_middleware_phone_return)):
    """
    Generates analytics for all groups managed by the authenticated user within a date range.
    
    Args:
        start_date (str): Start date in YYYY-MM-DD format
        end_date (str): End date in YYYY-MM-DD format
        phone (str): Authenticated user's phone (from dependency)
        
    Returns:
        dict: Complex analytics structure with completion statistics per group
        
    Example Request:
        GET /get_my_groups_analytic?start_date=2023-01-01&end_date=2023-01-31
        
    Example Response:
        {
            "Group1": [
                10,  // total tasks
                ["user1", 8, 1, 80, 20],  // user stats: [completed, late, % done, % not done]
                ["user2", 5, 2, 50, 50],
                ["Загалом", 10, 7, 3, 70, 30]  // group totals
            ],
            "Group2": [...]
        }
    """
    try:
        users_group4 = await groups.find(
            {"active": 1},
            {"_id": 0, "group_name": 1, "user_phones": 1}
        ).to_list(length=None)
        users_group = {item["group_name"]: item["user_phones"] for item in users_group4}
        groups_user2 = dict()

        pipeline = [
            {
                "$match": {
                    "created_by": phone,
                    "end_date": {  
                        "$gte": start_date,  
                        "$lte": end_date  
                    }
                }
            },
            {
                "$group": {
                    "_id": "$group",  
                    "total_tasks": { "$sum": 1 }
                }
            },
            {
                "$sort": { "total_tasks": -1 }  
            }
        ]

        results = await tasks.aggregate(pipeline).to_list(length=None)
        
        groups_user = dict()
        for doc in results:
            e = list()
            e.append(doc['total_tasks'])
            groups_user[doc['_id']] = e
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
        end_date = datetime.strptime(end_date, "%Y-%m-%d")
        start_date = datetime.combine(start_date, datetime.min.time())  
        end_date = datetime.combine(end_date, datetime.max.time())
        
        pipeline = [
            {
                "$match": {
                    "status": 1,  
                }
            },
            {
                "$addFields": {
                    "finish_time_parsed": {
                        "$dateFromString": {
                            "dateString": "$finish_time",  
                            "format": "%d.%m.%Y, %H:%M:%S"  
                        }
                    }
                }
            },
            {
                "$match": {
                    "finish_time_parsed": {
                        "$gte": start_date,  
                        "$lt": end_date      
                    }
                }
            },
            {
                "$group": {
                    "_id": {
                        "group": "$group",  
                        "phone": "$phone"   
                    },
                    "total_count": { "$sum": 1 },  
                    "in_time_1_count": {
                        "$sum": {
                            "$cond": [{ "$eq": ["$in_time", 1] }, 1, 0]  
                        }
                    },
                    "in_time_0_count": {
                        "$sum": {
                            "$cond": [{ "$eq": ["$in_time", 0] }, 1, 0]  
                        }
                    }
                }
            },
            {
                "$sort": {
                    "_id.group": 1,  
                    "_id.phone": 1   
                }
            }
        ]
        
        total_count_group = dict()
        total_completed_tasks = dict()
        
        results = await completedtasks.aggregate(pipeline).to_list(length=None)
        
        for result in results:
            d = list()
            group = result["_id"]["group"]
            phone = result["_id"]["phone"]
            total_count = result["total_count"]
            total_count_group[group] = total_count
            in_time_1_count = result["in_time_1_count"]
            in_time_0_count = result["in_time_0_count"]
            if group not in total_completed_tasks:
                total_completed_tasks[group] = dict()
            if 'completed_tasks' not in total_completed_tasks[group]:
                total_completed_tasks[group]['completed_tasks'] = 0
            if 'not_in_time' not in total_completed_tasks[group]:
                total_completed_tasks[group]['not_in_time'] = 0
            total_completed_tasks[group]['completed_tasks'] += total_count
            total_completed_tasks[group]['not_in_time'] += in_time_0_count
            d.append(phone)
            d.append(total_count)
            d.append(in_time_0_count)
            d.append(int((total_count / groups_user[group][0]) * 100))
            d.append(int((1 - total_count / groups_user[group][0]) * 100)) 
            groups_user[group].append(d)         
            
            if group not in groups_user2:
                groups_user2[group] = []
            groups_user2[group].append(phone)
        
        for group in groups_user.keys():
            all_phones = set(users_group[group])
            if group in groups_user2:
                selected_phones = set(groups_user2[group])
            else:
                selected_phones = set()
            diff = all_phones - selected_phones
            for phone in diff:
                d = list()
                d.append(phone)
                d.append(0)
                d.append(0)
                d.append(0)
                d.append(100) 
                groups_user[group].append(d)
        for group in groups_user.keys():
            total_users = len(users_group[group]) * groups_user[group][0]
            if group in total_completed_tasks:
                total_completed_task = total_completed_tasks[group]['completed_tasks']
                total_not_in_time = total_completed_tasks[group]['not_in_time']
                compl_procent = int(total_completed_tasks[group]['completed_tasks'] / ((len(users_group[group]) * groups_user[group][0])) * 100)
                puncompl_procent = 100 - compl_procent
            else:
                total_completed_task = 0
                total_not_in_time = 0
                compl_procent = 0
                puncompl_procent = 100 
            groups_user[group].append(["Загалом", total_users, total_completed_task, total_not_in_time, compl_procent, puncompl_procent])
        
        return groups_user
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.post("/download_excel_tasks_analytic")
async def download_excel(start_date: str, end_date: str, group: str, groups_data: list = Body(...),  phone=Depends(auth_middleware_phone_return)):
    """
    Generates and downloads an Excel report for a specific group's task analytics.
    
    Args:
        start_date (str): Report start date
        end_date (str): Report end date
        group (str): Group name for the report
        groups_data (list): Pre-calculated analytics data
        phone (str): Authenticated user's phone
        
    Returns:
        StreamingResponse: Excel file download
        
    Example Request:
        POST /download_excel_tasks_analytic
        start_date=2023-01-01&end_date=2023-01-31&group=Development
        Body: [pre-calculated analytics data]
        
    Response:
        Excel file download with name "tasks_report.xlsx"
    """
    try:
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Виконані завдання"
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = f"{start_date} - {end_date}"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.append([])  
        ws['A3'] = group
        ws['B3'] = "Задачі"
        ws.merge_cells("B3:D3")  
        ws['E3'] = "Результат"
        ws.merge_cells("E3:F3")  
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        border_style = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        for col in ['A3', 'B3', 'E3']:
            cell = ws[col]
            cell.fill = green_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style

        data_header = ["Користувач", "Поставлено", "Виконано", "Невчасно", "% Виконання", "% Невиконання"]
        ws.append(data_header)  
        for col_num, cell_value in enumerate(data_header, start=1):
            cell = ws.cell(row=4, column=col_num)
            cell.fill = green_fill
            cell.font = Font(bold=True)
            cell.border = border_style

        current_row = 5

        for group in groups_data[1:-1]:
            k = list()
            k.append(group[0])  
            k.append(groups_data[0]) 
            k.extend(group[1:])  
            ws.append(k)
            current_row += 1  
        ws.append(groups_data[-1])
        last_row = ws.max_row
        for col in range(1, len(groups_data[-1]) + 1):
            cell = ws.cell(row=last_row, column=col)
            cell.fill = green_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=tasks_report.xlsx"}
        )
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.post("/download_excel_tasks_analytic2")
async def download_excel_multiple_groups(
    start_date: str,
    end_date: str,
    groups: list = Body(...),
    groups_data: dict = Body(...),
    phone=Depends(auth_middleware_phone_return)
):
    """
    Generates and downloads Excel report for multiple groups' analytics.
    
    Args:
        start_date (str): Report start date
        end_date (str): Report end date
        groups (list): Group names to include
        groups_data (dict): Pre-calculated analytics data
        phone (str): Authenticated user's phone
        
    Returns:
        StreamingResponse: Excel file download
        
    Example Request:
        POST /download_excel_tasks_analytic2
        start_date=2023-01-01&end_date=2023-01-31
        Body: {
            "groups": ["Development", "Design"],
            "groups_data": {pre-calculated data...}
        }
        
    Response:
        Excel file download with name "tasks_report_multiple.xlsx"
    """
    try:
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Звіт по задачах"
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = f"{start_date} - {end_date}"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row = 3  
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        light_blue_fill = PatternFill(start_color="FFE1F5FE", end_color="FFE1F5FE", fill_type="solid")
        border_style = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        for group in groups:
            group_data = groups_data.get(group, [])
            if not group_data:
                continue
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            cell = ws.cell(row=current_row, column=1)
            cell.value = group
            cell.fill = light_blue_fill
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in range(1, 7):  
                border_cell = ws.cell(row=current_row, column=col)
                border_cell.border = border_style
            current_row += 1
            header1 = ["Користувач", "Поставлено", "Виконано", "Невчасно", "% Виконання", "% Невиконання"]
            for col_num, header in enumerate(header1, start=1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = green_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_style
            current_row += 1
            for entry in group_data[1:-1]:
                row_data = [entry[0], group_data[0], entry[1], entry[2], entry[3], entry[4]]
                for col_num, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = value
                    cell.border = border_style
                current_row += 1
            last_entry = group_data[-1]
            last_row_data = [last_entry[0], last_entry[1], last_entry[2], last_entry[3], last_entry[4], last_entry[5]]

            for col_num, value in enumerate(last_row_data, start=1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.fill = green_fill  
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_style
            current_row += 1

        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=tasks_report_multiple.xlsx"}
        )
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")
    
@user_app.get("/get_my_groups")
async def login_user(request: Request, phone = Depends(auth_middleware_phone_return)):
    """
    Retrieves group names where the user is manager.
    
    Args:
        phone (str): Authenticated user's phone (from dependency)
        
    Returns:
        list: Group names managed by user
        
    Example Request:
        GET /get_my_groups
        
    Example Response:
        ["Development", "Design"]
    """
    try:
        results = await groups.find({"manager_phone": phone}, {"group_name": 1, "_id": 0}).to_list(length=None)
        results2 = [i["group_name"] for i in results]
        return results2
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.get("/get_my_info")
async def get_info(request: Request, phone = Depends(auth_middleware_phone_return)):
    """
    Retrieves authenticated user's information (excluding password).
    
    Args:
        phone (str): Authenticated user's phone (from dependency)
        
    Returns:
        dict: User information
        
    Example Request:
        GET /get_my_info
        
    Example Response:
        {
            "name": "John Doe",
            "phone": "+380123456789",
            "status": "manager",
            "telegramName": "johndoe"
        }
    """
    try:
        results = await users_collections.find({"phone": phone}, {"password": 0, "_id": 0}).to_list(length=None)
        re2 = list(results)
        return jsonable_encoder(re2)
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.post("/tasks")
async def create_task(request: Request, task: Task, phone=Depends(auth_middleware_phone_return)):
    """
    Creates one or multiple tasks based on task type (general, weekly, or single).
    
    Args:
        task (Task): Task creation data
        phone (str): Authenticated user's phone (from dependency)
        
    Returns:
        dict: Success message with count of created tasks
        
    Raises:
        HTTPException: 404 if user/group not found, 500 for server errors
        
    Example Request:
        POST /tasks
        {
            "title": "Weekly Meeting",
            "description": "Team sync",
            "startDate": "2023-01-01",
            "endDate": "2023-12-31",
            "startTime": "10:00",
            "endTime": "11:00",
            "repeatDays": ["Monday", "Wednesday"],
            "group": "Development",
            "taskType": "weekly",
            "importance": 2,
            "needphoto": 1,
            "needcomment": 0
        }
        
    Example Response:
        {"message": "104 tasks successfully saved to database"}
    """
    try:
        result = await groups.find_one({"group_name": task.group}, {"manager_phone": 1, "_id": 0})
        user_info = await users_collections.find_one({'phone': phone}, {'name': 1})

        if not result or result["manager_phone"] != phone:
            raise HTTPException(status_code=404, detail="У вас немає прав для виконання цієї задачі")

        if not user_info:
            raise HTTPException(status_code=404, detail="Користувача не знайдено")

        try:
            task_type = task.taskType
            start_date = datetime.strptime(task.startDate, "%Y-%m-%d")
            end_date = datetime.strptime(task.endDate, "%Y-%m-%d")

            days_to_create = []

            if task_type == "general":
                current = start_date
                while current <= end_date:
                    days_to_create.append(current)
                    current += timedelta(days=1)

            elif task_type == "weekly":
                repeat_days = task.repeatDays  
                weekday_indices = [list(calendar.day_name).index(day) for day in repeat_days]

                current = start_date
                while current <= end_date:
                    if current.weekday() in weekday_indices:
                        days_to_create.append(current)
                    current += timedelta(days=1)
            else:
                days_to_create = [start_date]

            for day in days_to_create:
                task_data = {
                    "title": task.title,
                    "description": task.description,
                    "start_date": day.strftime("%Y-%m-%d"),
                    "end_date": day.strftime("%Y-%m-%d"),
                    "start_time": task.startTime,
                    "end_time": task.endTime,
                    "repeat_days": task.repeatDays,
                    "group": task.group,
                    "task_type": task.taskType,
                    "importance": int(task.importance),
                    "created_by": phone,
                    'needphoto': task.needphoto,
                    'needcomment': task.needcomment,
                    'openquestion': task.openquestion,
                    "created_name": user_info['name']
                }
                await tasks.insert_one(task_data)

            return {"message": f"{len(days_to_create)} tasks successfully saved to database"}

        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to save task: {str(e)}")
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.get("/get_my_task")
async def get_tasks(request: Request, phone=Depends(auth_middleware_phone_return)):
    """
    Retrieves tasks assigned to authenticated user's groups.
    
    Args:
        phone (str): Authenticated user's phone (from dependency)
        
    Returns:
        list: Tasks sorted by importance with completed task IDs appended
        
    Example Request:
        GET /get_my_task
        
    Example Response:
        [
            {task1 data...},
            {task2 data...},
            ["completed_task_id1", "completed_task_id2"]
        ]
    """
    try:
        r = await groups.find({'user_phones': f"{phone}", 'active': 1}, {"group_name": 1, "_id": 0}).to_list(length=None)
        compltasks = await completedtasks.find({"phone": f"{phone}"}, {"id_task": 1, "_id": 0}).to_list(length=None)
        
        tasksCompleteIDs = [task['id_task'] for task in compltasks]
        
        groups_name = [group["group_name"] for group in r]
        tasks_cursor = tasks.find({'group': {'$in': groups_name}}).sort([('importance', -1)])
        user_tasks = await tasks_cursor.to_list(length=None)

        for task in user_tasks:
            task["_id"] = str(task["_id"])

        user_tasks.append(tasksCompleteIDs)
        return user_tasks
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.post("/push_task")
async def push_task(
    start_time: str = Form(...),
    finish_time: str = Form(...),
    pause_start: str = Form(...),
    pause_end: str = Form(...),
    id_task: str = Form(...),
    group: str = Form(...),
    task_name: str = Form(...),
    keyTime: str = Form(...),
    comment: str = Form(None),
    in_time: int = Form(...),
    images: List[UploadFile] = File([]),
    phone=Depends(auth_middleware_phone_return)
):
    """
    Records completion of a task with time tracking and optional photos.
    
    Args (form data):
        start_time: Task start time (DD.MM.YYYY, HH:MM:SS)
        finish_time: Task end time (DD.MM.YYYY, HH:MM:SS)
        pause_start: JSON list of pause start times
        pause_end: JSON list of pause end times
        id_task: Original task ID
        group: Task group name
        task_name: Task name
        keyTime: Unique time key
        comment: Optional completion comment
        in_time: 1 if completed on time, 0 otherwise
        images: List of uploaded photos
        
    Returns:
        dict: Success message
        
    Example Request:
        POST /push_task
        Form data:
        start_time=01.01.2023,10:00:00
        finish_time=01.01.2023,11:30:00
        pause_start=["01.01.2023,10:30:00"]
        pause_end=["01.01.2023,10:45:00"]
        id_task=507f1f77bcf86cd799439011
        group=Development
        task_name=Code Review
        keyTime=unique123
        comment=Completed successfully
        in_time=1
        images=[photo1.jpg, photo2.jpg]
        
    Example Response:
        {"message": "Informations about task successfully saved to database"}
    """
    try:
        pause_start_list = json.loads(pause_start)
        pause_end_list = json.loads(pause_end)
        try:
            dt_start = datetime.strptime(start_time, "%d.%m.%Y, %H:%M:%S")
            dt_finish = datetime.strptime(finish_time, "%d.%m.%Y, %H:%M:%S")
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат времени")
        total_minutes = (dt_finish - dt_start).total_seconds() / 60
        total_pause_minutes = 0
        for start_str, end_str in zip(pause_start_list, pause_end_list):
            try:
                ps = datetime.strptime(start_str, "%d.%m.%Y, %H:%M:%S")
                pe = datetime.strptime(end_str, "%d.%m.%Y, %H:%M:%S")
                total_pause_minutes += (pe - ps).total_seconds() / 60
            except Exception:
                continue

        active_minutes = total_minutes - total_pause_minutes
        photo_refs = []
        for image in images:
            content = await image.read()
            file_id = await fs.upload_from_stream(image.filename, content)
            photo_refs.append({
                "file_id": str(file_id),
                "filename": image.filename
            })
        task_data = {
            "start_time": start_time,
            "finish_time": finish_time,
            "pause_start": pause_start_list,
            "pause_end": pause_end_list,
            "id_task": id_task,
            "group": group,
            "keyTime": keyTime,
            "phone": phone,
            "comment": comment,
            "in_time": in_time,
            "task_name": task_name,
            "status": 1,
            "photos": photo_refs,
            "total_minutes": int(total_minutes),
            "pause_minutes": int(total_pause_minutes),
            "active_minutes": int(active_minutes)
        }

        await completedtasks.insert_one(task_data)

        return {"message": "Informations about task successfully saved to database"}
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")
        
@user_app.post("/cancel_task")
async def cancel_task(request: Request, task_cancel: TaskTimeCancel, phone=Depends(auth_middleware_phone_return)):
    """
    Records task cancellation with reason.
    
    Args:
        task_cancel (TaskTimeCancel): Cancellation details
        phone (str): Authenticated user's phone
        
    Returns:
        dict: Success message
        
    Example Request:
        POST /cancel_task
        {
            "cancel_time": "01.01.2023,12:00:00",
            "id_task": "507f1f77bcf86cd799439011",
            "task_name": "Code Review",
            "group": "Development",
            "comment": "Not enough time"
        }
        
    Example Response:
        {"message": "Informations about task successfully saved to database"}
    """
    try:
        task_data = {
            "finish_time": task_cancel.cancel_time,
            "id_task": task_cancel.id_task,
            "phone": phone,
            "comment": task_cancel.comment,
            "group": task_cancel.group,
            "task_name": task_cancel.task_name,
            'status': 0
        }
        await completedtasks.insert_one(task_data)
        return {"message": "Informations about task successfully saved to database"}
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.get("/get_my_created_task/")
async def get_created_tasks(request: Request, phone=Depends(auth_middleware_phone_return)):
    """
    Retrieves tasks created by the authenticated user.
    
    Args:
        phone (str): Authenticated user's phone (from dependency)
        
    Returns:
        list: Tasks sorted by importance
        
    Example Request:
        GET /get_my_created_task/
        
    Example Response:
        [
            {
                "_id": "507f1f77bcf86cd799439011",
                "title": "Code Review",
                "description": "Review PRs",
                ...
            },
            ...
        ]
    """
    try:
        tasks_cursor = tasks.find({'created_by': phone}).sort([('importance', -1)])
        user_tasks = await tasks_cursor.to_list(length=None)
        for task in user_tasks:
            task["_id"] = str(task["_id"])
        return user_tasks
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.get("/get_infoprocent_about_task/{group}/{task_id}")
async def get_info_procent_about_task(request: Request, group: str, task_id: str):
    """
    Calculates completion percentage for a specific task in a group.
    
    Args:
        group (str): Group name
        task_id (str): Task ID to check
        
    Returns:
        float: Completion percentage (0-100)
        
    Example Request:
        GET /get_infoprocent_about_task/Development/507f1f77bcf86cd799439011
        
    Example Response:
        75.0  // 75% completion rate
    """
    try:
        task_id = unquote(task_id)
        count = await groups.find_one({'group_name': group})
        count2 = await completedtasks.find({'id_task': task_id, 'status': 1}).to_list(length=None)
        return (len(count2)/len(count['user_phones'])) * 100
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")
    
@user_app.delete("/delete_task/{task_id}")
async def delete_task(request: Request, task_id: str, phone=Depends(auth_middleware_phone_return)):
    
    """
    Deletes a task created by the authenticated user.
    
    Args:
        task_id (str): Task ID to delete
        phone (str): Authenticated user's phone (must be task creator)
        
    Returns:
        dict: Success message
        
    Raises:
        HTTPException: 404 if task not found or not creator
        
    Example Request:
        DELETE /delete_task/507f1f77bcf86cd799439011
        
    Example Response:
        {"message": "Task successfully deleted"}
    """
    try:
        result = await tasks.delete_one({"_id": ObjectId(task_id), 'created_by': phone})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="The task was not found or you do not have sufficient rights")
        
        return {"message": "Task successfully deleted"}
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail="Помилка бази даних")

    except Exception as e:
        raise HTTPException(status_code=500, detail="Невідома помилка сервера")

@user_app.post("/question-task")
async def question_task(
    data: QuestionTaskRequest,
    phone=Depends(auth_middleware_phone_return)
):
    message_doc = {
        "task_id": data.taskId,
        "task_title": data.taskTitle,
        "text": data.comment,
        "created_at": datetime.utcnow(),
        "author": {
            "phone": phone,
            "role": "client"
        },
        "receiver": {
            "phone": data.createdBy,
            "name": data.createdName
        },
        "type": "question"
    }

    await comments.insert_one(message_doc)

    return {
        "status": "ok",
        "message": "Дані отримано",
        "taskId": data.taskId
    }

@user_app.get("/chats")
async def get_chats(phone: str = Depends(auth_middleware_phone_return)):
    pipeline = [
        {"$match": {"$or": [{"author.phone": phone}, {"receiver.phone": phone}]}},
        {"$group": {
            "_id": "$task_id",
            "task_id": {"$first": "$task_id"},
            "task_title": {"$first": "$task_title"},
            "author": {"$first": "$author"},
            "receiver": {"$first": "$receiver"}
        }},
        {"$sort": {"task_id": -1}}
    ]
    chats_raw = await comments.aggregate(pipeline).to_list(length=None)

    chats = []
    for chat in chats_raw:
        other_user_phone = (
            chat["receiver"]["phone"]
            if chat["author"]["phone"] == phone
            else chat["author"]["phone"]
        )
        
        other_user = await users_collections.find_one(
            {"phone": other_user_phone},
            {"password": 0}  
        )
        
        if other_user:
            other_user["_id"] = str(other_user["_id"])  

        chats.append({
            "task_id": chat["task_id"],
            "task_title": chat["task_title"],
            "other_user": other_user
        })

    return {"status": "ok", "chats": chats}


@user_app.get("/messages/{task_id}/{other_phone}")
async def get_chat_messages(task_id: str, other_phone: str, my_phone: str = Depends(auth_middleware_phone_return)):
    cursor = comments.find({
        "task_id": task_id,
        "$or": [
            {"author.phone": my_phone, "receiver.phone": other_phone},
            {"author.phone": other_phone, "receiver.phone": my_phone}
        ]
    }).sort("created_at", 1)

    messages = []
    async for msg in cursor:
        msg["_id"] = str(msg["_id"])
        messages.append(msg)

    return {"status": "ok", "messages": messages}

connections: Dict[str, Dict[str, WebSocket]] = {}

@user_app.websocket("/ws/chat/{task_id}/{phones_pair}")
async def chat_ws(websocket: WebSocket, task_id: str, phones_pair: str):
    # Декодируем URL
    phones_pair = unquote(phones_pair)

    token = websocket.query_params.get("token")
    if not token:
        await websocket.close(code=1008)
        return

    try:
        payload = jwt.decode(token, "test", algorithms=["HS256"])
        my_phone = payload.get("sub")
    except:
        await websocket.close(code=1008)
        return

    await websocket.accept()

    # Используем ровно тот chat_id, который прислали с фронта
    chat_id = f"{task_id}_{phones_pair}"

    if chat_id not in connections:
        connections[chat_id] = {}
    connections[chat_id][my_phone] = websocket
    print(f"[CONNECT] chat_id={chat_id}, phones={list(connections[chat_id].keys())}")

    try:
        while True:
            data = await websocket.receive_json()

            message_doc = {
                "task_id": task_id,
                "task_title": data["task_title"],
                "text": data["text"],
                "created_at": datetime.utcnow(),
                "author": {"phone": my_phone, "role": "client"},
                "receiver": data["receiver"],
                "type": "question"
            }

            await comments.insert_one(message_doc)
            outgoing_message = dict(message_doc)
            outgoing_message["created_at"] = message_doc["created_at"].isoformat()

            if "_id" in outgoing_message:
                outgoing_message["_id"] = str(outgoing_message["_id"])

            # Рассылаем всем подключенным по chat_id
            for phone, ws in connections[chat_id].items():
                try:
                    await ws.send_json(outgoing_message)
                    print(f"[SENT] to {phone}: {outgoing_message['text']}")
                except Exception as e:
                    print(f"[ERROR] sending to {phone}: {e}")
            print(f"[BROADCAST] chat_id={chat_id}, phones={list(connections[chat_id].keys())}")
            receiver_phone = data["receiver"]["phone"]
            ws_global = global_connections.get(receiver_phone)

            if ws_global:
                await ws_global.send_json({
                    "type": "new_message",
                    "task_id": task_id,
                    "from_phone": my_phone,
                    "created_at": outgoing_message["created_at"]
                })
            receiver_in_chat = (
                chat_id in connections and
                receiver_phone in connections[chat_id]
            )

            if receiver_in_chat:
                print("❌ Пользователь в чате — Telegram не отправляем")
            else:
                print("✅ Пользователь НЕ в чате — отправляем в Telegram")
                user = await users_collections.find_one({"phone": receiver_phone})
                if not user:
                    print("Пользователь с таким телефоном не найден")
                else:
                    username = user.get("telegramName") 
                    print(f"Найден пользователь: {username}") 

                    telegram_user = await telegram_users.find_one({"username": username}) 
                    if not telegram_user:
                        print("Chat_id для Telegram не найден")
                    else:
                        chat_id_tg = telegram_user.get("chat_id")
                        print(f"Chat_id Telegram: {chat_id_tg}")

                try:
                    await bot.send_message(
                        chat_id=chat_id_tg,
                        text=f"Нове повідомлення по таску '{data['task_title']}' від {my_phone}:\n{data['text']}"
                    )
                except Exception as e:
                    print(f"TG error: {e}")

    except WebSocketDisconnect:
        if chat_id in connections and my_phone in connections[chat_id]:
            del connections[chat_id][my_phone]
            if not connections[chat_id]:
                del connections[chat_id]

global_connections: Dict[str, WebSocket] = {}

@user_app.websocket("/ws/notifications")
async def notifications_ws(websocket: WebSocket):
    token = websocket.query_params.get("token")
    if not token:
        await websocket.close(code=1008)
        return

    try:
        payload = jwt.decode(token, "test", algorithms=["HS256"])
        my_phone = payload.get("sub")
    except:
        await websocket.close(code=1008)
        return

    await websocket.accept()
    global_connections[my_phone] = websocket
    print(f"[GLOBAL CONNECT] phone={my_phone}")

    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        global_connections.pop(my_phone, None)
        print(f"[GLOBAL DISCONNECT] phone={my_phone}")
            
@user_app.post("/chats/read")
async def mark_chat_read(
    body: ChatReadRequest,
    my_phone: str = Depends(auth_middleware_phone_return)
):
    await chat_read_state.update_one(
        {
            "user_phone": my_phone,
            "task_id": body.task_id,
            "other_user_phone": body.other_user_phone
        },
        {
            "$set": {
                "last_read_at": datetime.utcnow()
            }
        },
        upsert=True
    )

    return {"status": "ok"}

@user_app.get("/chats/unread")
async def get_unread_chats(
    phone: str = Depends(auth_middleware_phone_return)
):
    pipeline = [
    {
        "$match": {"receiver.phone": phone}
    },
    {
        "$lookup": {
            "from": "ChatReadState",
            "let": {"task_id": "$task_id", "author_phone": "$author.phone"},
            "pipeline": [
                {
                    "$match": {
                        "$expr": {
                            "$and": [
                                {"$eq": ["$user_phone", phone]},
                                {"$eq": ["$task_id", "$$task_id"]},
                                {"$eq": ["$other_user_phone", "$$author_phone"]}
                            ]
                        }
                    }
                }
            ],
            "as": "read_state"
        }
    },
    {
        "$unwind": {
            "path": "$read_state",
            "preserveNullAndEmptyArrays": True  
        }
    },
    {
        "$group": {
            "_id": {"task_id": "$task_id", "from_phone": "$author.phone"},
            "last_message_at": {"$max": "$created_at"},
            "last_read_at": {"$max": "$read_state.last_read_at"} 
        }
    },
    {
        "$project": {
            "task_id": "$_id.task_id",
            "from_phone": "$_id.from_phone",
            "last_message_at": 1,
            "last_read_at": 1,
            "is_unread": {
                "$cond": [
                    {"$or": [{"$eq": ["$last_read_at", None]}, {"$gt": ["$last_message_at", "$last_read_at"]}]},
                    True,
                    False
                ]
            }
        }
    }
]

    unread = await comments.aggregate(pipeline).to_list(None)

    return {
        "status": "ok",
        "debug": unread, 
        "unread": [
            f"{u['task_id']}_{u['from_phone']}" for u in unread if u["is_unread"]
        ]
    }
    
@user_app.put("/update_task/")
async def update_task(request: Request, task: TaskEdit, phone=Depends(auth_middleware_phone_return)):
    """
    Updates an existing task created by the authenticated user.
    
    Args:
        task (TaskEdit): Updated task data
        phone (str): Authenticated user's phone (must be task creator)
        
    Returns:
        dict: Success message
        
    Raises:
        HTTPException: 404 if task not found or not creator
        
    Example Request:
        PUT /update_task/
        {
            "title": "Updated Task",
            "description": "New description",
            "start_date": "2023-01-01",
            "end_date": "2023-01-01",
            "start_time": "09:00",
            "end_time": "10:00",
            "repeat_days": [],
            "group": "Development",
            "task_type": "single",
            "importance": 2,
            "created_by": "+380123456789",
            "taskid": "507f1f77bcf86cd799439011",
            "needphoto": 1,
            "needcomment": 1
        }
        
    Example Response:
        {"message": "Task successfully updated"}
    """
    task_data = {
        "title": task.title,
        "description": task.description,
        "start_date": task.start_date,
        "end_date": task.end_date,
        "start_time": task.start_time,
        "end_time": task.end_time,
        "repeat_days": task.repeat_days,
        "group": task.group,
        "task_type": task.task_type,
        "importance": task.importance,
        'needcomment': task.needcomment,
        'needphoto': task.needphoto,
        'openquestion': task.openquestion,
    }
    try:
        result = await tasks.update_one(
            {"_id": ObjectId(task.taskid), 'created_by': phone},
            {"$set": task_data}
        )
        
        if result.modified_count > 0:
            return {"message": "Task successfully updated"}
        else:
            raise HTTPException(status_code=404, detail="Failed to update task")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update task: {str(e)}")

@user_app.get("/tasks/{task_id}")
async def get_task_by_id(
    task_id: str,
    phone=Depends(auth_middleware_phone_return)
):
    if not ObjectId.is_valid(task_id):
        raise HTTPException(status_code=400, detail="Invalid task id")

    task = await tasks.find_one(
        {"_id": ObjectId(task_id)},
        {"_id": 0}
    )

    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    return {
        "status": "ok",
        "task": task
    }